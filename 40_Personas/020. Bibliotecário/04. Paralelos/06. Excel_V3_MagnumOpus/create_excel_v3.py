import os
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Configuração
csv_dir = r"40_Personas\020. Bibliotecário\03. Paralelos\06. Excel_V3_MagnumOpus\02. Dados_Estruturados_CSV"
output_file = r"40_Personas\020. Bibliotecário\03. Paralelos\06. Excel_V3_MagnumOpus\Excel_V3_MagnumOpus.xlsx"

# Estilos
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Azul Profissional
border_style = Side(border_style="thin", color="000000")
thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
alignment_center = Alignment(horizontal="center", vertical="center")
alignment_left = Alignment(horizontal="left", vertical="center")

def format_worksheet(ws):
    # Formatar Cabeçalho
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    # Ajustar largura das colunas e formatar células de dados
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        
        for cell in col:
            # Aplicar borda em todas as células
            cell.border = thin_border
            if cell.row > 1: # Dados (não cabeçalho)
                cell.alignment = alignment_left
            
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Ajuste fino da largura (limite máximo para não ficar gigante)
        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > 50: adjusted_width = 50 # Max width cap
        ws.column_dimensions[column].width = adjusted_width

    # Congelar Paineis (Primeira Linha e Primeira Coluna)
    ws.freeze_panes = "B2"

print("Iniciando a construção do Magnum Opus... 🏛️")

# Criar Workbook
wb = Workbook()
# Remover a aba padrão "Sheet"
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# Listar CSVs ordenados
csv_files = sorted([f for f in os.listdir(csv_dir) if f.endswith(".csv")])

for csv_file in csv_files:
    tab_name = csv_file.replace(".csv", "").replace("01_", "").replace("02_", "").replace("03_", "").replace("04_", "").replace("05_", "").replace("06_", "").replace("07_", "").replace("08_", "").replace("09_", "")
    # Limitar tamanho do nome da aba (Excel limita a 31 chars)
    tab_name = tab_name[:30]
    
    print(f"Adicionando aba: {tab_name}...")
    
    # Ler CSV com Pandas (mais robusto)
    csv_path = os.path.join(csv_dir, csv_file)
    df = pd.read_csv(csv_path)
    
    # Criar aba
    ws = wb.create_sheet(title=tab_name)
    
    # Escrever dados do DataFrame na aba
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
        
    # Aplicar formatação
    format_worksheet(ws)

# Salvar
wb.save(output_file)
print(f"✅ Excel Impecável gerado com sucesso: {output_file}")
